# PATH: src/app/generar_excel_crosstabs_completo.py

from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import datetime

from src.service.generar_crosstab_modelo_materiales import cargar_datos, preparar_datos_coois, generar_crosstab_modelo_materiales
from src.service.formato_crosstab import format_crosstabs, agregar_cantidad_bom_header, formato_indice

def generar_excel_crosstabs_completo(archivo, sheet_bom_ea, sheet_bom_eb, sheet_coois):
    print('Cargando datos...')
    bom_ea, bom_eb, coois = cargar_datos(archivo, sheet_bom_ea, sheet_bom_eb, sheet_coois)
    coois_ea, coois_eb = preparar_datos_coois(coois)

    results = []
    for bom, coois, subset_name in [(bom_ea, coois_ea, 'EA'), (bom_eb, coois_eb, 'EB')]:
        print(f'Generando archivo Excel para {subset_name}...')
        wb = Workbook()
        wb.remove(wb.active)  # Elimina la pestaña predeterminada

        index_sheet = wb.create_sheet(title="Índice", index=0)
        index_sheet.append(["Modelo"])
        unique_modelos = sorted(bom['Modelo'].unique())

        for modelo in unique_modelos:
            ws = wb.create_sheet(title=modelo)
            crosstab = generar_crosstab_modelo_materiales(bom, coois, modelo)
            for row in dataframe_to_rows(crosstab, index=True, header=True):
                ws.append(row)
            agregar_cantidad_bom_header(ws, bom, modelo)
            format_crosstabs(ws, bom, modelo)

        formato_indice(index_sheet)
        timestamp = datetime.datetime.now().strftime("%Y%m%d%H%M%S")
        filename = f'./results/crosstabs_{subset_name}/crosstabs_materiales_{subset_name}_{timestamp}.xlsx'
        wb.save(filename)
        results.append(filename)

    return results